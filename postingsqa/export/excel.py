"""Build the Excel workbook: Dashboard (charts), Jobs, Rejected, QA Summary, Raw."""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from postingsqa.config import SOURCE_NAMES
from postingsqa.export import charts
from postingsqa.models import Job, RunSummary
from postingsqa.qa.checks import annualize
from postingsqa.qa.pipeline import QAReport

HEADER_FILL = PatternFill("solid", fgColor="1f2933")
HEADER_FONT = Font(bold=True, color="ffffff")
KPI_FONT = Font(bold=True, size=20, color="1f2933")
KPI_LABEL_FONT = Font(size=10, color="52514e")
TITLE_FONT = Font(bold=True, size=16, color="0b0b0b")
NEW_FILL = PatternFill("solid", fgColor="e6f2ff")
THIN = Side(style="thin", color=charts.GRID)

JOB_COLUMNS: list[tuple[str, str, int]] = [
    # (header, attribute, width)
    ("Source", "source", 11),
    ("Title", "title", 38),
    ("Company", "company", 26),
    ("Location", "location", 24),
    ("Remote", "remote", 9),
    ("Posted", "posted_at", 11),
    ("Salary Min", "salary_min", 12),
    ("Salary Max", "salary_max", 12),
    ("Period", "salary_period", 8),
    ("Currency", "salary_currency", 9),
    ("Estimate", "salary_is_estimate", 9),
    ("Type", "employment_type", 12),
    ("Seniority", "seniority", 14),
    ("New", "is_new", 7),
    ("First Seen", "first_seen", 12),
    ("Query", "search_query", 18),
    ("URL", "url", 40),
    ("Description", "description", 60),
]

SALARY_BUCKETS = [
    ("< $50k", 0, 50_000),
    ("$50–75k", 50_000, 75_000),
    ("$75–100k", 75_000, 100_000),
    ("$100–125k", 100_000, 125_000),
    ("$125–150k", 125_000, 150_000),
    ("$150–200k", 150_000, 200_000),
    ("$200k+", 200_000, float("inf")),
]


def _cell_value(job: Job, attr: str):
    v = getattr(job, attr)
    if attr == "description" and v:
        return v[:2000]
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return v


def _write_header(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")


def _write_jobs_sheet(ws: Worksheet, jobs: list[Job], extra: list[tuple[str, int]] | None = None, reasons: list[str] | None = None, table_name: str = "Jobs") -> None:
    headers = [h for h, _, _ in JOB_COLUMNS] + [h for h, _ in (extra or [])]
    widths = [w for _, _, w in JOB_COLUMNS] + [w for _, w in (extra or [])]
    _write_header(ws, headers)
    for r, job in enumerate(jobs, start=2):
        for c, (_, attr, _) in enumerate(JOB_COLUMNS, 1):
            cell = ws.cell(row=r, column=c, value=_cell_value(job, attr))
            if attr == "url" and job.url:
                cell.hyperlink = job.url
                cell.style = "Hyperlink"
            elif attr == "posted_at" and job.posted_at:
                cell.number_format = "yyyy-mm-dd"
            elif attr == "first_seen" and job.first_seen:
                cell.number_format = "yyyy-mm-dd"
            elif attr in ("salary_min", "salary_max") and cell.value is not None:
                cell.number_format = "#,##0"
            elif attr == "description":
                cell.alignment = Alignment(wrap_text=False)
        if reasons is not None:
            ws.cell(row=r, column=len(JOB_COLUMNS) + 1, value=reasons[r - 2])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    last_col = get_column_letter(len(headers))
    last_row = max(len(jobs) + 1, 2)
    table = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(table)
    if jobs and table_name == "Jobs":
        new_col = get_column_letter([a for _, a, _ in JOB_COLUMNS].index("is_new") + 1)
        ws.conditional_formatting.add(
            f"A2:{last_col}{last_row}",
            FormulaRule(formula=[f'${new_col}2="Yes"'], fill=NEW_FILL),
        )


class _DataTables:
    """Writes small helper tables to the hidden `_data` sheet and remembers their ranges."""

    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.col = 1

    def write(self, header: tuple[str, str], rows: list[tuple[object, float]]) -> tuple[int, int, int, int]:
        """Returns (min_row, max_row, cat_col, val_col)."""
        cat_col, val_col = self.col, self.col + 1
        self.ws.cell(row=1, column=cat_col, value=header[0])
        self.ws.cell(row=1, column=val_col, value=header[1])
        for i, (k, v) in enumerate(rows, start=2):
            c = self.ws.cell(row=i, column=cat_col, value=k)
            if isinstance(k, date):
                c.number_format = "mmm d"
            self.ws.cell(row=i, column=val_col, value=v)
        self.col += 3
        return 1, max(len(rows) + 1, 2), cat_col, val_col


def _salary_midpoint_yearly(job: Job) -> float | None:
    vals = [v for v in (job.salary_min, job.salary_max) if v is not None]
    if not vals or (job.salary_currency and job.salary_currency.upper() != "USD"):
        return None
    return annualize(sum(vals) / len(vals), job.salary_period)


def _postings_per_day(jobs: list[Job], days: int = 30) -> list[tuple[str, float]]:
    counts = Counter(j.posted_at for j in jobs if j.posted_at)
    if not counts:
        return []
    end = date.today()
    start = max(min(counts), end - timedelta(days=days - 1))
    return [(start + timedelta(days=i), counts.get(start + timedelta(days=i), 0)) for i in range((end - start).days + 1)]


def _write_dashboard(ws: Worksheet, data: Worksheet, kept: list[Job], report: QAReport, summary: RunSummary | None) -> None:
    ws.sheet_view.showGridLines = False
    # Print / PDF export: the whole dashboard on one landscape page.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws["A1"] = "Job Postings Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = KPI_LABEL_FONT
    from postingsqa.sources import attribution_text

    credit = attribution_text(report.per_source)
    if credit:
        ws["A3"] = credit
        ws["A3"].font = KPI_LABEL_FONT

    new_count = sum(1 for j in kept if j.is_new)
    blocked = ", ".join(summary.blocked_sources) if summary and summary.blocked_sources else "none"
    kpis = [
        ("Scraped", report.scraped),
        ("Passed QA", len(report.kept)),
        ("Rejected", len(report.rejected)),
        ("New this run", new_count),
        ("Blocked sources", blocked),
    ]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 3
        ws.cell(row=4, column=col, value=value).font = KPI_FONT
        ws.cell(row=5, column=col, value=label).font = KPI_LABEL_FONT
    for i in range(1, 30):
        ws.column_dimensions[get_column_letter(i)].width = 9

    tables = _DataTables(data)
    anchors = iter(["A7", "J7", "A23", "J23", "A39", "J39", "A55"])

    # 1. Jobs by source (identity → fixed categorical color per source)
    by_source = Counter(j.source for j in kept)
    order = [s for s in SOURCE_NAMES if s in by_source] + sorted(s for s in by_source if s not in SOURCE_NAMES)
    rows = [(s.capitalize(), by_source[s]) for s in order]
    if rows:
        rng = tables.write(("Source", "Jobs"), rows)
        ws.add_chart(charts.bar_chart(data, "Jobs passing QA by source", *rng, point_colors=[charts.SOURCE_COLOR.get(s, charts.NEUTRAL) for s in order], x_title="Source"), next(anchors))

    # 2. Top companies (magnitude → single hue, sorted)
    top = Counter(j.company for j in kept if j.company).most_common(10)
    if top:
        rng = tables.write(("Company", "Jobs"), [(c, n) for c, n in top])
        ws.add_chart(charts.bar_chart(data, "Top 10 companies", *rng, horizontal=True, x_title="Company"), next(anchors))

    # 3. Postings per day
    per_day = _postings_per_day(kept)
    if per_day:
        rng = tables.write(("Date", "Jobs"), per_day)
        ws.add_chart(charts.line_chart(data, "Postings per day (last 30 days)", *rng), next(anchors))

    # 4. Remote vs on-site
    remote = Counter("Remote" if j.remote else ("On-site / hybrid" if j.remote is not None else "Unknown") for j in kept)
    if remote:
        rows = [(k, remote[k]) for k in ("Remote", "On-site / hybrid", "Unknown") if k in remote]
        rng = tables.write(("Work mode", "Jobs"), rows)
        ws.add_chart(charts.pie_chart(data, "Remote vs on-site", *rng, colors=[charts.CATEGORICAL[2], charts.CATEGORICAL[3], charts.NEUTRAL]), next(anchors))

    # 5. Salary distribution (bucketed yearly USD midpoints)
    buckets = Counter()
    for j in kept:
        mid = _salary_midpoint_yearly(j)
        if mid is None:
            continue
        for label, lo, hi in SALARY_BUCKETS:
            if lo <= mid < hi:
                buckets[label] += 1
                break
    if buckets:
        rows = [(label, buckets.get(label, 0)) for label, _, _ in SALARY_BUCKETS]
        rng = tables.write(("Salary (USD/yr)", "Jobs"), rows)
        ws.add_chart(charts.bar_chart(data, "Salary distribution (jobs with salary data)", *rng, color=charts.CATEGORICAL[6], x_title="Yearly salary"), next(anchors))

    # 6. QA rejection reasons
    if report.rejection_counts:
        rows = sorted(report.rejection_counts.items(), key=lambda kv: -kv[1])
        rng = tables.write(("Check", "Rejections"), rows)
        ws.add_chart(charts.bar_chart(data, "QA rejections by check", *rng, horizontal=True, color=charts.STATUS_BAD, x_title="Check", y_title="Rejected"), next(anchors))

    # 7. New vs previously seen
    if kept:
        rows = [("New", new_count), ("Seen before", len(kept) - new_count)]
        rng = tables.write(("Status", "Jobs"), rows)
        ws.add_chart(charts.pie_chart(data, "New vs previously seen", *rng, colors=[charts.CATEGORICAL[0], charts.NEUTRAL]), next(anchors))


def _write_qa_summary(ws: Worksheet, report: QAReport, summary: RunSummary | None) -> None:
    ws["A1"] = "QA Summary"
    ws["A1"].font = TITLE_FONT
    _write_header(ws, ["Source", "Scraped", "Kept", "Rejected", "Pass rate", "Status"], row=3)
    r = 4
    present = set(report.per_source) | (set(summary.blocked_sources) | set(summary.errors) if summary else set())
    for src in [s for s in SOURCE_NAMES if s in present] + sorted(present - set(SOURCE_NAMES)):
        stats = report.per_source.get(src)
        blocked = summary is not None and src in summary.blocked_sources
        error = summary.errors.get(src) if summary else None
        scraped = stats["scraped"] if stats else 0
        kept = stats["kept"] if stats else 0
        ws.cell(row=r, column=1, value=src.capitalize())
        ws.cell(row=r, column=2, value=scraped)
        ws.cell(row=r, column=3, value=kept)
        ws.cell(row=r, column=4, value=scraped - kept)
        rate = ws.cell(row=r, column=5, value=(kept / scraped) if scraped else None)
        rate.number_format = "0%"
        ws.cell(row=r, column=6, value="blocked" if blocked else (f"error: {error}" if error else ("ok" if stats else "no data")))
        r += 1

    r += 1
    _write_header(ws, ["Check", "Rejections", "Flags (soft warnings)"], row=r)
    r += 1
    from postingsqa.qa.checks import CHECKS

    for name in ["duplicate"] + [n for n, _ in CHECKS]:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=report.rejection_counts.get(name, 0))
        ws.cell(row=r, column=3, value=report.flag_counts.get(name, 0))
        r += 1
    for col, w in zip("ABCDEF", (22, 12, 22, 12, 10, 30)):
        ws.column_dimensions[col].width = w


def build_workbook(report: QAReport, all_jobs: list[Job], summary: RunSummary | None, out_path: Path) -> Path:
    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    jobs_ws = wb.create_sheet("Jobs")
    rejected_ws = wb.create_sheet("Rejected")
    qa_ws = wb.create_sheet("QA Summary")
    raw_ws = wb.create_sheet("Raw")
    data_ws = wb.create_sheet("_data")
    data_ws.sheet_state = "hidden"

    # new jobs first, then newest posting date first
    kept = sorted(report.kept, key=lambda j: (not j.is_new, -(j.posted_at or date.min).toordinal()))
    _write_jobs_sheet(jobs_ws, kept, table_name="Jobs")
    _write_jobs_sheet(rejected_ws, [r.job for r in report.rejected], extra=[("Reason", 50)], reasons=[r.reason for r in report.rejected], table_name="Rejected")
    _write_qa_summary(qa_ws, report, summary)
    _write_jobs_sheet(raw_ws, all_jobs, table_name="Raw")
    _write_dashboard(dashboard, data_ws, kept, report, summary)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
