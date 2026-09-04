import random
from datetime import date, timedelta, datetime, timezone

from openpyxl import load_workbook

from postingsqa.config import QAConfig
from postingsqa.export.excel import build_workbook
from postingsqa.models import Job, RunSummary
from postingsqa.qa.pipeline import run_qa


def synthetic_jobs(n=40):
    rng = random.Random(1)
    titles = ["QA Engineer", "Data Analyst", "Automation Engineer", "SDET", "Senior QA Engineer", "Forklift Operator"]
    companies = ["Acme", "Beta Inc", "Gamma LLC", "Delta Staffing", "Epsilon"]
    jobs = []
    for i in range(n):
        src = rng.choice(["linkedin", "indeed", "glassdoor"])
        host = {"linkedin": "www.linkedin.com/jobs/view", "indeed": "www.indeed.com/viewjob?jk=", "glassdoor": "www.glassdoor.com/job-listing"}[src]
        lo = rng.choice([None, 60000, 90000, 120000, 40])
        jobs.append(Job(
            source=src, source_id=str(i), title=rng.choice(titles), company=rng.choice(companies),
            location=rng.choice(["Remote", "Austin, TX", "New York, NY", "Berlin"]),
            url=f"https://{host}{i}", posted_at=date.today() - timedelta(days=rng.randint(0, 40)),
            salary_min=lo, salary_max=(lo * 1.3 if lo else None), salary_period="hour" if lo == 40 else "year",
            description="Lorem ipsum " * rng.choice([5, 40]), is_new=rng.random() < 0.5,
            first_seen=datetime.now(timezone.utc),
        ))
    return jobs


def test_workbook_has_sheets_and_charts(tmp_path):
    jobs = synthetic_jobs()
    report = run_qa(jobs, QAConfig())
    assert report.kept and report.rejected
    summary = RunSummary(run_id="t", started_at=datetime.now(timezone.utc), blocked_sources=["indeed"], errors={"glassdoor": "timeout"})
    out = build_workbook(report, jobs, summary, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == ["Dashboard", "Jobs", "Rejected", "QA Summary", "Raw", "_data"]
    assert wb["_data"].sheet_state == "hidden"
    assert len(wb["Dashboard"]._charts) >= 6
    jobs_ws = wb["Jobs"]
    assert jobs_ws.max_row == len(report.kept) + 1
    assert jobs_ws["A1"].value == "Source" and "Jobs" in jobs_ws.tables
    url_col = [c.value for c in jobs_ws[1]].index("URL") + 1
    assert jobs_ws.cell(row=2, column=url_col).hyperlink is not None
    rej = wb["Rejected"]
    assert rej.cell(row=1, column=rej.max_column).value == "Reason"
    assert rej.cell(row=2, column=rej.max_column).value
    assert wb["Raw"].max_row == len(jobs) + 1
    qa = wb["QA Summary"]
    assert qa["F5"].value == "blocked"  # indeed row
